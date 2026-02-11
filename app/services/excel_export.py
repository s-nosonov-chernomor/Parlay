# app/services/excel_export.py
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def _excel_safe(v):
    # Excel/openpyxl не поддерживает tz-aware datetimes
    if isinstance(v, datetime):
        if v.tzinfo is not None:
            return v.replace(tzinfo=None)
        return v
    return v

def build_xlsx(rows: list[dict], meta: dict) -> bytes:
    wb = Workbook()

    # ---- sheet: meta ----
    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_meta.append(["key", "value"])
    for k, v in meta.items():
        ws_meta.append([k, str(v)])

    # ---- sheet: data ----
    ws = wb.create_sheet("data")
    columns = ["ts", "ui_id", "zone_code", "source_id", "bind_key", "note", "topic", "value_num", "value_text"]
    ws.append(columns)

    for r in rows:
        ws.append([
            _excel_safe(r.get("ts")),
            r.get("ui_id"),
            r.get("zone_code"),
            r.get("source_id"),
            r.get("bind_key"),
            r.get("note"),
            r.get("topic"),
            r.get("value_num"),
            r.get("value_text"),
        ])

    # простая авто-ширина
    for i, col in enumerate(columns, start=1):
        max_len = len(col)
        for cell in ws.iter_cols(min_col=i, max_col=i, min_row=1, max_row=min(ws.max_row, 2000)):
            for c in cell:
                if c.value is None:
                    continue
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, max_len + 2))

    # ---- sheet: pivots (удобно под графики фронту/аналитику) ----
    # Формат: ts | <note/bind_key@ui_id>...
    ws_p = wb.create_sheet("pivot")
    # группируем: (ui_id, bind_key) -> {ts -> value_num/value_text}
    series = {}
    all_ts = set()
    for r in rows:
        key = f"{r.get('ui_id')}::{r.get('bind_key')}::{(r.get('note') or '')}".strip()
        series.setdefault(key, {})
        ts = _excel_safe(r.get("ts"))
        all_ts.add(ts)
        val = r.get("value_num")
        if val is None:
            val = r.get("value_text")
        series[key][ts] = val

    all_ts_sorted = sorted([t for t in all_ts if t is not None])
    keys_sorted = sorted(series.keys())

    ws_p.append(["ts"] + keys_sorted)
    for ts in all_ts_sorted:
        row = [_excel_safe(ts)]
        for k in keys_sorted:
            row.append(series[k].get(ts))
        ws_p.append(row)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
